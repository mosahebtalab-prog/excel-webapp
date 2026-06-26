<<<<<<< HEAD
import os
import openpyxl
from openpyxl import load_workbook
from flask import Flask, render_template, request, redirect, url_for
import jdatetime

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'

# ========== تنظیم مسیر فایل اکسل ==========
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, 'data.xlsx')

# ========== توابع مدیریت فایل اکسل ==========
def ensure_excel_file():
    """اگر فایل اکسل وجود نداشته باشد، با ساختار کامل می‌سازد"""
=======
from flask import Flask, render_template, request, redirect, url_for, session
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime
import json
import jdatetime

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# مسیر فایل اکسل
EXCEL_FILE = 'data.xlsx'

# تابع برای اطمینان از وجود فایل اکسل و ایجاد آن در صورت نیاز
def ensure_excel_file():
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        
        # برگه بیماران
        ws_patients = wb.active
        ws_patients.title = 'patients'
        ws_patients.append(['id', 'name', 'phone', 'disease', 'created_at'])
        
<<<<<<< HEAD
        # برگه ویزیت‌ها
        ws_visits = wb.create_sheet('visits')
        ws_visits.append(['id', 'patient_id', 'visit_date', 'reason', 'treatment', 'cost', 'created_at'])
        
        wb.save(EXCEL_FILE)
        print(f"✅ فایل اکسل در مسیر {EXCEL_FILE} ساخته شد.")

def get_all_patients():
    """دریافت همه بیماران از اکسل"""
    ensure_excel_file()
=======
        # برگه ویزیت‌ها (با ستون جدید treatment)
        ws_visits = wb.create_sheet('visits')
        ws_visits.append(['id', 'patient_id', 'visit_date', 'reason', 'treatment', 'cost', 'created_at'])
        
        # برگه درآمد
        ws_income = wb.create_sheet('income')
        ws_income.append(['id', 'patient_id', 'amount', 'date', 'description'])
        
        wb.save(EXCEL_FILE)

# تابع برای دریافت تمام بیماران
def get_all_patients():
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
    wb = load_workbook(EXCEL_FILE)
    ws = wb['patients']
    patients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            patients.append({
                'id': row[0],
                'name': row[1] or '',
                'phone': row[2] or '',
                'disease': row[3] or '',
                'created_at': row[4] or ''
            })
    wb.close()
    return patients

<<<<<<< HEAD
def get_patient_by_id(patient_id):
    """دریافت یک بیمار با ID"""
    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb['patients']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == int(patient_id):
            wb.close()
            return {
                'id': row[0],
                'name': row[1] or '',
                'phone': row[2] or '',
                'disease': row[3] or '',
                'created_at': row[4] or ''
            }
    wb.close()
    return None

def get_patient_visits(patient_id):
    """دریافت ویزیت‌های یک بیمار"""
    ensure_excel_file()
=======
# تابع برای دریافت ویزیت‌های یک بیمار
def get_patient_visits(patient_id):
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
    wb = load_workbook(EXCEL_FILE)
    ws = wb['visits']
    visits = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] == int(patient_id):
            visits.append({
                'id': row[0],
                'patient_id': row[1],
                'visit_date': row[2] or '',
                'reason': row[3] or '',
                'treatment': row[4] or '',
                'cost': row[5] or 0,
                'created_at': row[6] or ''
            })
    wb.close()
    return visits

<<<<<<< HEAD
def get_total_income():
    """محاسبه درآمد کل از همه ویزیت‌ها"""
    ensure_excel_file()
=======
# تابع برای دریافت اطلاعات یک بیمار
def get_patient_by_id(patient_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb['patients']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == int(patient_id):
            wb.close()
            return {
                'id': row[0],
                'name': row[1] or '',
                'phone': row[2] or '',
                'disease': row[3] or '',
                'created_at': row[4] or ''
            }
    wb.close()
    return None

# تابع برای محاسبه درآمد کل
def get_total_income():
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
    wb = load_workbook(EXCEL_FILE)
    ws = wb['visits']
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[5] is not None:
            try:
                total += int(row[5])
            except:
                pass
    wb.close()
    return total

<<<<<<< HEAD
def get_income_by_date_range(start_date, end_date):
    """محاسبه درآمد در بازه زمانی شمسی"""
=======
# تابع برای محاسبه درآمد در بازه زمانی
def get_income_by_date_range(start_date, end_date):
    wb = load_workbook(EXCEL_FILE)
    ws = wb['visits']
    total = 0
    visits = []
    
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
    try:
        start_jd = jdatetime.datetime.strptime(start_date, '%Y/%m/%d')
        end_jd = jdatetime.datetime.strptime(end_date, '%Y/%m/%d')
    except:
        return 0, []
<<<<<<< HEAD

    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb['visits']
    total_income = 0
    visits_in_range = []

=======
    
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[2] is not None:
            try:
                visit_date = row[2]
                if '/' in visit_date:
                    visit_jd = jdatetime.datetime.strptime(visit_date, '%Y/%m/%d')
                    if start_jd <= visit_jd <= end_jd:
                        cost = int(row[5]) if row[5] else 0
<<<<<<< HEAD
                        total_income += cost
                        visits_in_range.append({
=======
                        total += cost
                        visits.append({
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
                            'id': row[0],
                            'patient_id': row[1],
                            'visit_date': visit_date,
                            'reason': row[3] or '',
                            'treatment': row[4] or '',
                            'cost': cost
                        })
            except:
                pass
    
    wb.close()
<<<<<<< HEAD
    return total_income, visits_in_range

# ========== مسیرها (Routes) ==========
@app.route('/')
def index():
=======
    return total, visits

# مسیر اصلی
@app.route('/')
def index():
    ensure_excel_file()
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
    patients = get_all_patients()
    total_income = get_total_income()
    return render_template('index.html', patients=patients, total_income=total_income)

<<<<<<< HEAD
=======
# مسیر ثبت بیمار جدید
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
@app.route('/add_patient', methods=['POST'])
def add_patient():
    name = request.form.get('name', '').strip()
    phone = request.form.get('phone', '').strip()
    disease = request.form.get('disease', '').strip()
<<<<<<< HEAD

    if not name:
        return redirect(url_for('index'))

    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb['patients']

=======
    
    if not name:
        return redirect(url_for('index'))
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb['patients']
    
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
    # پیدا کردن آخرین ID
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[0] > max_id:
            max_id = row[0]
<<<<<<< HEAD

    new_id = max_id + 1
    now = jdatetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    ws.append([new_id, name, phone, disease, now])
    wb.save(EXCEL_FILE)
    wb.close()
    return redirect(url_for('index'))

@app.route('/delete_patient/<int:patient_id>')
def delete_patient(patient_id):
    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb['patients']
=======
    
    new_id = max_id + 1
    now = jdatetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    
    ws.append([new_id, name, phone, disease, now])
    wb.save(EXCEL_FILE)
    wb.close()
    
    return redirect(url_for('index'))

# مسیر حذف بیمار
@app.route('/delete_patient/<int:patient_id>')
def delete_patient(patient_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb['patients']
    
    # پیدا کردن ردیف بیمار
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
    row_to_delete = None
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == patient_id:
            row_to_delete = idx
            break
<<<<<<< HEAD
    if row_to_delete:
        ws.delete_rows(row_to_delete)
        wb.save(EXCEL_FILE)
    wb.close()
    return redirect(url_for('index'))

=======
    
    if row_to_delete:
        ws.delete_rows(row_to_delete)
    
    wb.save(EXCEL_FILE)
    wb.close()
    
    return redirect(url_for('index'))

# مسیر پروفایل بیمار
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
@app.route('/profile/<int:patient_id>')
def profile(patient_id):
    patient = get_patient_by_id(patient_id)
    if not patient:
        return redirect(url_for('index'))
<<<<<<< HEAD
    visits = get_patient_visits(patient_id)
    return render_template('profile.html', patient=patient, visits=visits)

=======
    
    visits = get_patient_visits(patient_id)
    return render_template('profile.html', patient=patient, visits=visits)

# مسیر ثبت ویزیت جدید (با فیلد treatment)
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
@app.route('/add_visit', methods=['POST'])
def add_visit():
    patient_id = request.form.get('patient_id', '').strip()
    visit_date = request.form.get('visit_date', '').strip()
    reason = request.form.get('reason', '').strip()
    treatment = request.form.get('treatment', '').strip()
    cost = request.form.get('cost', '').strip()
<<<<<<< HEAD

    if not patient_id or not visit_date:
        return redirect(url_for('index'))

    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb['visits']

    # پیدا کردن آخرین ID برای ویزیت
=======
    
    if not patient_id or not visit_date:
        return redirect(url_for('index'))
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb['visits']
    
    # پیدا کردن آخرین ID
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[0] > max_id:
            max_id = row[0]
<<<<<<< HEAD

    new_id = max_id + 1
    now = jdatetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    ws.append([new_id, int(patient_id), visit_date, reason, treatment, int(cost) if cost else 0, now])
    wb.save(EXCEL_FILE)
    wb.close()
    return redirect(url_for('profile', patient_id=patient_id))

@app.route('/delete_visit/<int:visit_id>/<int:patient_id>')
def delete_visit(visit_id, patient_id):
    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb['visits']
=======
    
    new_id = max_id + 1
    now = jdatetime.datetime.now().strftime('%Y/%m/%d %H:%M')
    
    ws.append([new_id, int(patient_id), visit_date, reason, treatment, int(cost) if cost else 0, now])
    wb.save(EXCEL_FILE)
    wb.close()
    
    return redirect(url_for('profile', patient_id=patient_id))

# مسیر حذف ویزیت
@app.route('/delete_visit/<int:visit_id>/<int:patient_id>')
def delete_visit(visit_id, patient_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb['visits']
    
    # پیدا کردن ردیف ویزیت
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
    row_to_delete = None
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == visit_id:
            row_to_delete = idx
            break
<<<<<<< HEAD
    if row_to_delete:
        ws.delete_rows(row_to_delete)
        wb.save(EXCEL_FILE)
    wb.close()
    return redirect(url_for('profile', patient_id=patient_id))

=======
    
    if row_to_delete:
        ws.delete_rows(row_to_delete)
    
    wb.save(EXCEL_FILE)
    wb.close()
    
    return redirect(url_for('profile', patient_id=patient_id))

# مسیر گزارش درآمد
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
@app.route('/income_report', methods=['GET', 'POST'])
def income_report():
    total_income = 0
    visits = []
    start_date = ''
    end_date = ''
<<<<<<< HEAD

    if request.method == 'POST':
        start_date = request.form.get('start_date', '').strip()
        end_date = request.form.get('end_date', '').strip()

        if start_date and end_date:
            total_income, visits = get_income_by_date_range(start_date, end_date)

    return render_template('income_report.html',
                         total_income=total_income,
                         visits=visits,
                         start_date=start_date,
                         end_date=end_date)

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)
=======
    
    if request.method == 'POST':
        start_date = request.form.get('start_date', '').strip()
        end_date = request.form.get('end_date', '').strip()
        
        if start_date and end_date:
            total_income, visits = get_income_by_date_range(start_date, end_date)
    
    return render_template('income_report.html', 
                         total_income=total_income, 
                         visits=visits, 
                         start_date=start_date, 
                         end_date=end_date)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
>>>>>>> fbea6f1e53eae5b252d4c87a3680ba82a0aab7fb
